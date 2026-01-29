"""
Generate SFTP Specification Excel file for Natera FIND -> Five9
"""
import subprocess
import sys

# Install openpyxl if not available
subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Create workbook
wb = Workbook()

# Define styles
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
tbd_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def style_header_row(ws, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

def auto_width(ws):
    for column_cells in ws.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        length = min(length + 2, 50)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length

def add_borders(ws):
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                cell.border = thin_border

def highlight_tbd(ws):
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and "[TBD]" in str(cell.value):
                cell.fill = tbd_fill

# =============================================================================
# TAB 1: Overview
# =============================================================================
ws1 = wb.active
ws1.title = "Overview"

overview_data = [
    ["Property", "Sandbox", "Production", "Notes"],
    ["Document Title", "SFTP Delivery Specification: Natera FIND → Five9", "", ""],
    ["Version", "1.0", "", ""],
    ["Last Updated", "[TBD]", "", ""],
    ["", "", "", ""],
    ["CONNECTION DETAILS", "", "", ""],
    ["SFTP Host", "[TBD]", "[TBD]", "Required"],
    ["Port", "22", "22", "Standard SFTP port"],
    ["S3 Path", "s3://sftp-walgreens-sandbox/walgreens/OUT/XXXX/", "s3://sftp-walgreens/walgreens/OUT/XXXX/", "Confirm XXXX placeholder"],
    ["Authentication Method", "[TBD - SSH Key / Password / AWS]", "[TBD]", "Required"],
    ["Username", "[TBD]", "[TBD]", "Required"],
    ["IP Allowlist (Source)", "[TBD]", "[TBD]", "Curebase outbound IPs"],
    ["", "", "", ""],
    ["SCHEDULE", "", "", ""],
    ["Frequency", "Daily", "Daily", ""],
    ["Delivery Time", "[TBD]", "[TBD]", "Required - specify time"],
    ["Timezone", "[TBD]", "[TBD]", "Required - e.g., America/New_York"],
    ["Weekend Handling", "Fri-Sun data delivered Monday", "Fri-Sun data delivered Monday", "No weekend processing"],
    ["Empty File Policy", "No file if 0 participants", "No file if 0 participants", ""],
    ["", "", "", ""],
    ["FILE FORMAT", "", "", ""],
    ["Filename Pattern", "NATERA_FIND_stuck_in_journey_YYYY-MM-DD.csv.pgp", "", ""],
    ["Format", "CSV (PGP encrypted)", "", ""],
    ["Delimiter", "[TBD - comma assumed]", "", "Required"],
    ["Encoding", "[TBD - UTF-8 / ASCII]", "", "Required"],
    ["Header Row", "[TBD - Yes / No]", "", "Required"],
    ["Line Ending", "[TBD - CRLF / LF]", "", "Required"],
    ["", "", "", ""],
    ["PGP ENCRYPTION", "", "", ""],
    ["Public Key", "[TBD - attach or reference]", "", "Required"],
    ["Algorithm", "[TBD]", "", ""],
]

for row_idx, row_data in enumerate(overview_data, 1):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws1.cell(row=row_idx, column=col_idx, value=value)
        if row_idx == 1:
            cell.font = header_font
            cell.fill = header_fill
        if value in ["CONNECTION DETAILS", "SCHEDULE", "FILE FORMAT", "PGP ENCRYPTION"]:
            cell.font = Font(bold=True, size=11)

style_header_row(ws1, 4)
add_borders(ws1)
highlight_tbd(ws1)
auto_width(ws1)

# =============================================================================
# TAB 2: Field Definitions
# =============================================================================
ws2 = wb.create_sheet("Field Definitions")

field_data = [
    ["#", "Field Name", "Data Type", "Required", "Max Length", "Format / Valid Values", "Description"],
    [1, "REASON_FOR_CALL", "String", "Yes", "[TBD]", "Needs Consent | Incomplete Survey | Missing Colonoscopy", "Call disposition reason - determines call script"],
    [2, "CUREBASE_ID", "String", "Yes", "[TBD]", "[TBD - format unclear]", "Curebase participant unique identifier"],
    [3, "CUSTOM_ID", "String", "No", "[TBD]", "02-XXXX or null", "Custom identifier if available, else null"],
    [4, "PROTOCOL_ID", "String", "Yes", "4", "FIND", "Always 'FIND' for this protocol"],
    [5, "FIRST_NAME", "String", "Yes", "[TBD]", "", "Participant first name"],
    [6, "LAST_NAME", "String", "Yes", "[TBD]", "", "Participant last name"],
    [7, "PRIMARY_PHONE", "String", "Yes", "10", "10 digits, no formatting", "Primary contact phone number"],
    [8, "PATIENT_LANGUAGE", "String", "Yes", "1", "E = English, S = Spanish", "Preferred language for call"],
    [9, "DOB", "Date", "Yes", "10", "YYYY-MM-DD", "Participant date of birth"],
    [10, "ADDRESS", "String", "Yes", "[TBD]", "", "Street address line 1"],
    [11, "ADDRESS2", "String", "No", "[TBD]", "", "Street address line 2 (apt, suite, etc.)"],
    [12, "CITY", "String", "Yes", "[TBD]", "", "City name"],
    [13, "STATE", "String", "Yes", "2", "2-letter abbreviation (e.g., NY, CA)", "State code"],
    [14, "ZIP_CODE", "String", "Yes", "5", "5 digits", "Postal/ZIP code"],
]

for row_idx, row_data in enumerate(field_data, 1):
    for col_idx, value in enumerate(row_data, 1):
        ws2.cell(row=row_idx, column=col_idx, value=value)

style_header_row(ws2, 7)
add_borders(ws2)
highlight_tbd(ws2)
auto_width(ws2)

# =============================================================================
# TAB 3: Business Rules
# =============================================================================
ws3 = wb.create_sheet("Business Rules")

rules_data = [
    ["Rule Type", "Rule Name", "Condition", "Details"],
    ["INCLUSION", "Needs Consent", "In INFORMED_CONSENT state > 24 hours", "User stuck in consent step"],
    ["INCLUSION", "Incomplete Survey", "In PRESCREENING state > 48 hours", "User stuck in prescreening"],
    ["INCLUSION", "Missing Colonoscopy", "COLONOSCOPY_DATE visit opened but not completed within 48 hours", "User needs follow-up for colonoscopy date"],
    ["", "", "", ""],
    ["EXCLUSION", "Invalid Phone", "Phone number ≠ 10 digits", "Cannot dial invalid numbers"],
    ["EXCLUSION", "Invalid ZIP", "ZIP code missing or ≠ 5 digits", "Required for state compliance"],
    ["EXCLUSION", "Invalid State", "State missing or ≠ 2 letters", "Required for state compliance/call time restrictions"],
    ["EXCLUSION", "Duplicate", "Participant previously appeared on any list", "Delta logic - only new participants"],
    ["EXCLUSION", "Withdrawn", "Status = Withdrawn", "No longer in study"],
    ["EXCLUSION", "Opted Out", "Status = Opted-out", "Declined participation"],
    ["EXCLUSION", "Failed Screening", "Status = Failed Screening", "Did not qualify"],
    ["EXCLUSION", "Failed Pre-screening", "Status = Failed Pre-screening", "Did not qualify"],
    ["", "", "", ""],
    ["TIMING", "Weekend Accumulation", "States entered Fri/Sat/Sun", "Included in Monday file"],
    ["TIMING", "Empty File Suppression", "0 participants meet criteria", "No file generated"],
]

for row_idx, row_data in enumerate(rules_data, 1):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws3.cell(row=row_idx, column=col_idx, value=value)
        if value == "INCLUSION":
            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif value == "EXCLUSION":
            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        elif value == "TIMING":
            cell.fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")

style_header_row(ws3, 4)
add_borders(ws3)
auto_width(ws3)

# =============================================================================
# TAB 4: Validation & Error Handling
# =============================================================================
ws4 = wb.create_sheet("Validation & Errors")

validation_data = [
    ["Category", "Property", "Value", "Notes"],
    ["File Validation", "Checksum", "[TBD - None / MD5 / SHA-256]", "For integrity verification"],
    ["File Validation", "Acknowledgment File", "[TBD - Yes / No]", "Does Five9 send confirmation?"],
    ["File Validation", "Max File Size", "[TBD]", "Any size limits?"],
    ["", "", "", ""],
    ["Error Handling", "Retry Policy", "[TBD]", "Auto-retry on failure?"],
    ["Error Handling", "Failure Notification Email", "[TBD]", "Who gets notified?"],
    ["Error Handling", "Failure Notification Method", "[TBD - Email / Slack / PagerDuty]", ""],
    ["Error Handling", "SLA for Resolution", "[TBD]", "Expected response time"],
    ["", "", "", ""],
    ["Data Validation", "Phone Validation", "Must be exactly 10 digits", "Performed by Curebase before send"],
    ["Data Validation", "ZIP Validation", "Must be exactly 5 digits", "Performed by Curebase before send"],
    ["Data Validation", "State Validation", "Must be exactly 2 letters", "Performed by Curebase before send"],
]

for row_idx, row_data in enumerate(validation_data, 1):
    for col_idx, value in enumerate(row_data, 1):
        ws4.cell(row=row_idx, column=col_idx, value=value)

style_header_row(ws4, 4)
add_borders(ws4)
highlight_tbd(ws4)
auto_width(ws4)

# =============================================================================
# TAB 5: Contacts
# =============================================================================
ws5 = wb.create_sheet("Contacts")

contacts_data = [
    ["Role", "Organization", "Name", "Email", "Phone", "Notes"],
    ["Technical Lead", "Curebase", "[TBD]", "[TBD]", "[TBD]", "Primary technical contact"],
    ["Technical Lead", "Five9", "[TBD]", "[TBD]", "[TBD]", "Primary technical contact"],
    ["Project Manager", "Curebase", "[TBD]", "[TBD]", "[TBD]", ""],
    ["Project Manager", "Five9/Walgreens", "[TBD]", "[TBD]", "[TBD]", ""],
    ["Escalation", "Curebase", "[TBD]", "[TBD]", "[TBD]", "For critical issues"],
    ["Escalation", "Five9/Walgreens", "[TBD]", "[TBD]", "[TBD]", "For critical issues"],
    ["On-Call Support", "[TBD]", "[TBD]", "[TBD]", "[TBD]", "After-hours contact"],
]

for row_idx, row_data in enumerate(contacts_data, 1):
    for col_idx, value in enumerate(row_data, 1):
        ws5.cell(row=row_idx, column=col_idx, value=value)

style_header_row(ws5, 6)
add_borders(ws5)
highlight_tbd(ws5)
auto_width(ws5)

# =============================================================================
# TAB 6: Changelog
# =============================================================================
ws6 = wb.create_sheet("Changelog")

changelog_data = [
    ["Version", "Date", "Author", "Change Description"],
    ["1.0", "[TBD]", "[TBD]", "Initial specification created from original Word document"],
    ["", "", "", ""],
    ["", "", "", ""],
]

for row_idx, row_data in enumerate(changelog_data, 1):
    for col_idx, value in enumerate(row_data, 1):
        ws6.cell(row=row_idx, column=col_idx, value=value)

style_header_row(ws6, 4)
add_borders(ws6)
highlight_tbd(ws6)
auto_width(ws6)

# =============================================================================
# TAB 7: Open Questions
# =============================================================================
ws7 = wb.create_sheet("Open Questions")

questions_data = [
    ["#", "Question", "Source", "Owner", "Status", "Resolution"],
    [1, "What is the actual SFTP host/path? (S3 paths provided may not be SFTP)", "Original Doc", "[TBD]", "Open", ""],
    [2, "What is the format for CUREBASE_ID?", "Original Doc (marked with ?)", "[TBD]", "Open", ""],
    [3, "What time are daily files delivered?", "Missing from spec", "[TBD]", "Open", ""],
    [4, "What timezone is used for 24hr/48hr triggers?", "Missing from spec", "[TBD]", "Open", ""],
    [5, "Is there a header row in the CSV?", "Missing from spec", "[TBD]", "Open", ""],
    [6, "What PGP key should be used for encryption?", "Missing from spec", "[TBD]", "Open", ""],
    [7, "Should validation exist on participant-facing side instead of exclusion?", "Original Doc comment", "[TBD]", "Open", "Per Natalia: No, state-specific call time requirements"],
    [8, "Updated SFTP path needed", "Original Doc line 97", "[TBD]", "Open", ""],
]

for row_idx, row_data in enumerate(questions_data, 1):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws7.cell(row=row_idx, column=col_idx, value=value)
        if row_idx > 1 and col_idx == 5:
            if value == "Open":
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            elif value == "Resolved":
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

style_header_row(ws7, 6)
add_borders(ws7)
auto_width(ws7)

# =============================================================================
# Save workbook
# =============================================================================
output_path = r"C:\Users\zach.fabiano\Projects\Five9\Natera_FIND_SFTP_Spec_Five9.xlsx"
wb.save(output_path)
print(f"✅ Specification saved to: {output_path}")
print(f"\nTabs created:")
for sheet in wb.sheetnames:
    print(f"  - {sheet}")
print(f"\n⚠️  Items marked [TBD] are highlighted in yellow and need to be filled in.")

